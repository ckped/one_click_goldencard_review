import os
import io
import json
import pandas as pd
import streamlit as st
import google.generativeai as genai
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# âœ… è®€å–ç’°å¢ƒè®Šæ•¸
from dotenv import load_dotenv
load_dotenv()
API_KEY = os.getenv("GENIMI_API_KEY")
if not API_KEY:
    st.error("âŒ æ‰¾ä¸åˆ°ç’°å¢ƒè®Šæ•¸ GENIMI_API_KEYï¼Œè«‹å…ˆè¨­å®šå¾Œé‡æ–°åŸ·è¡Œ")
else:
    genai.configure(api_key=API_KEY)

# âœ… æ¨¡å‹èˆ‡æ¬„ä½è¨­å®š
model = genai.GenerativeModel("gemini-2.5-pro")
COLUMN_ORDER = [
    "è‹±æ–‡åå­—ï¼‹è‹±æ–‡å§“æ°",
    "ä¸­æ–‡å§“å",
    "ç¾è·å…¬å¸",
    "ç¾è·è·ç¨±",
    "æ•™è‚²èƒŒæ™¯(å­¸æ ¡)",
    "æ•™è‚²èƒŒæ™¯(ç³»æ‰€)",
    "å·¥ä½œç¶“æ­·",
    "å°ˆæ¥­æŠ€è¡“/è­‰ç…§",
    "å°ˆé•·é ˜åŸŸ",
    "å¯©æŸ¥æ„è¦‹æˆ–å‚™æ³¨"
]

OCR_PROMPT = """
è§’è‰²ï¼šä½ æ˜¯ä¸€å€‹ç²¾æº–çš„è³‡æ–™åˆ†æäººå“¡èˆ‡å¯©æŸ¥äººå“¡ï¼Œå°ˆé–€å¾è¤‡é›œæ–‡ä»¶ä¸­æå–é—œéµè³‡è¨Šã€‚
ä»»å‹™ï¼šè«‹ä½ ä»”ç´°é–±è®€ç”³è«‹äººç¶œåˆè³‡æ–™ PDFï¼Œå®ƒåŒ…å«CVç°¡æ­·ã€èƒŒæ™¯è£œå……èªªæ˜ç­‰ï¼Œè³‡æ–™æœƒåŒæ™‚åŒ…å«ä¸­æ–‡ã€è‹±æ–‡æˆ–å…¶å¤–æ–‡ã€‚
å·¥ä½œç¶“æ­·åˆ¤æ–·åŸå‰‡ï¼š
1.å…ˆåˆ¤æ–·ç”³è«‹äººç¾è·æˆ–æ›¾ä»»çš„å…¬å¸æ˜¯å¦ã€Œæœ¬èº«å³ç‚ºè»Ÿé«”æˆ–è³‡è¨Šæœå‹™æ¥­å…¬å¸ã€ã€‚é€™è¡¨ç¤ºå…¬å¸çš„ä¸»ç‡Ÿæ¥­å‹™å¿…é ˆæ˜¯è»Ÿé«”é–‹ç™¼ã€ç³»çµ±æ•´åˆã€è³‡è¨Šæœå‹™ã€é›²ç«¯æœå‹™ç­‰èˆ‡è»Ÿé«”æˆ–è³‡è¨Šç§‘æŠ€ç›´æ¥ç›¸é—œçš„ç”¢æ¥­é¡åˆ¥ã€‚
2.å…¬å¸å¦‚éå±¬ä¸Šè¿°è»Ÿé«”æˆ–è³‡è¨Šæœå‹™æ¥­ï¼Œåˆ¤æ–·çš„é‡å¿ƒç‚ºç”³è«‹äººçš„ã€Œè·å‹™å…§å®¹ã€ï¼Œé ˆã€Œæ˜ç¢ºé¡¯ç¤ºå…¶ä¸»è¦å·¥ä½œæ¶‰åŠè»Ÿé«”é–‹ç™¼ã€è»Ÿé«”å·¥å…·æ‡‰ç”¨æˆ–å…·å‚™å°ˆæ¥­è»Ÿé«”æŠ€è¡“å°ˆé•·ã€ã€‚


è¼¸å‡ºæ ¼å¼è¦æ±‚ï¼š è«‹ä½ åªè¼¸å‡ºä¸€å€‹ JSON æ ¼å¼çš„å­—ä¸² (JSON string)ï¼Œå…¶ä¸­åŒ…å«æ‰€æœ‰æˆ‘è¦æ±‚æå–çš„æ¬„ä½åŠå…¶å°æ‡‰çš„å€¼ã€‚è«‹ç¢ºä¿æ¯å€‹æ¬„ä½åç¨±èˆ‡æˆ‘æœªä¾† Google Sheet ä¸­çš„æ¨™é¡Œå®Œå…¨ä¸€è‡´ã€‚
å¦‚æœæŸå€‹æ¬„ä½æ‰¾ä¸åˆ°è³‡è¨Šï¼Œè«‹å°‡å…¶å€¼è¨­å®šç‚º "N/A" (ä¸å¯ç”¨)ã€‚
ä»¥ä¸‹æ˜¯æˆ‘è¦æ±‚æå–çš„æ¬„ä½åç¨± (JSON Key) åŠå…¶é æœŸçš„å€¼é¡å‹/èªªæ˜ï¼š
[
{
"è‹±æ–‡åå­—ï¼‹è‹±æ–‡å§“æ°": "string",
"ä¸­æ–‡å§“å": "string",
"ç¾è·å…¬å¸: "string",
"ç¾è·è·ç¨±": "string",
â€œæ•™è‚²èƒŒæ™¯(å­¸æ ¡)": "stringï¼Œå› å­¸æ­·æ¶µè“‹å¤§å­¸ã€ç ”ç©¶æ‰€å’Œåšå£«ç­‰ï¼Œå¦‚æœ‰æä¾›ä¸åŒéšæ®µçš„å­¸æ­·è­‰æ˜æˆ–èªªæ˜ï¼Œé¸æ“‡æœ€é«˜å­¸æ­·çš„å­¸æ ¡ç‚ºä¸»",
â€œæ•™è‚²èƒŒæ™¯(ç³»æ‰€)": "string",è¼¸å‡ºæ ¼å¼ç‚º å­¸ä½_ç³»æ‰€åç¨±æˆ–å°ˆæ¥­é ˜åŸŸåç¨±â€,
"å·¥ä½œç¶“æ­·": "stringï¼Œè«‹è©³ç´°èªªæ˜æ¯ä»½å·¥ä½œç¶“é©—ï¼Œæ ¼å¼ç‚ºï¼šæ¯ä¸€ç­†è³‡æ–™æ˜¯ï¼Œè¼ƒæ—©å¹´/æœˆ~è¼ƒæ™šå¹´/æœˆ_ä»»è·ä¼æ¥­_ä»»è·è·ç¨±ï¼Œæ¯ç­†è³‡æ–™ä»¥å¹´ä»½æ–°èˆŠä¾†æ’åˆ—ï¼Œå¾æœ€æ–°å¹´ä»½é–‹å§‹æ’åˆ—åˆ°éå»è¼ƒèˆŠå¹´ä»½â€ã€‚
"å°ˆæ¥­æŠ€è¡“/è­‰ç…§": "string",åŒ…æ‹¬æœƒä½¿ç”¨çš„æ•¸ä½æŠ€èƒ½æˆ–å·¥å…·ï¼Œè­‰ç…§çš„æ ¼å¼ç‚ºï¼šå¹´åº¦ï¼¿è­‰ç…§åç¨±ï¼¿ç™¼è­‰æ©Ÿæ§‹â€,
"å°ˆé•·é ˜åŸŸ": "stringï¼Œä¾æ“šå·¥ä½œç¶“é©—é€²åˆ¤æ–·ã€‚â€
"å¯©æŸ¥æ„è¦‹æˆ–å‚™æ³¨": "stringï¼Œé‡å°ã€Œå·¥ä½œç¶“é©—ã€æ¯ä¸€ç­†è³‡æ–™ï¼Œä»¥100å­—æ‘˜è¦æ•´ç†å‡ºä»–åœ¨é€™å…¬å¸çš„å·¥ä½œå…§å®¹èˆ‡äº‹è¹Ÿæˆå°±ï¼Œä¸­æ–‡å›å¾©ã€‚ä¸¦è«‹æ ¹æ“šã€Œç”³è«‹è³‡æ ¼æ¢ä»¶ã€èˆ‡ã€Œå·¥ä½œç¶“æ­·åˆ¤æ–·åŸå‰‡ã€å°ç”³è«‹äººçš„å·¥ä½œç¶“æ­·é€²è¡Œç¶œåˆè©•ä¼°ã€‚é¦–å…ˆåˆ¤æ–·ç”³è«‹äººæ‰€å±¬å…¬å¸æ˜¯å¦ã€Œæœ¬èº«å³ç‚ºè»Ÿé«”æˆ–è³‡è¨Šæœå‹™æ¥­å…¬å¸ã€ã€‚è‹¥éï¼Œå‰‡åˆ¤æ–·ç”³è«‹äººã€Œè·å‹™å…§å®¹ã€æ˜¯å¦ã€Œæ˜ç¢ºé¡¯ç¤ºå…¶ä¸»è¦å·¥ä½œæ¶‰åŠè»Ÿé«”é–‹ç™¼ã€è»Ÿé«”å·¥å…·æ‡‰ç”¨æˆ–å…·å‚™å°ˆæ¥­è»Ÿé«”æŠ€è¡“å°ˆé•·ã€ã€‚è«‹è©³ç´°é—¡è¿°è©•ä¼°éç¨‹èˆ‡çµæœï¼ŒæŒ‡å‡ºç”³è«‹äººçš„å·¥ä½œç¶“é©—æ˜¯å¦ç¬¦åˆæ•¸ä½ç¶“æ¿Ÿç›¸é—œç”¢æ¥­æˆ–å°ˆæ¥­æŠ€è¡“è¦æ±‚ï¼Œè‹¥æœ‰ä¸ç¬¦æˆ–éœ€é€²ä¸€æ­¥é‡æ¸…ä¹‹è™•ï¼Œè«‹å…·é«”èªªæ˜ã€‚ä¾‹å¦‚ï¼Œè‹¥å…¬å¸éè»Ÿé«”æ¥­ï¼Œä¸”è·å‹™åå‘æ•¸ä½å·¥å…·ä½¿ç”¨è€…è€Œéé–‹ç™¼è€…ï¼Œå‰‡æ‡‰åœ¨å‚™è¨»ä¸­èªªæ˜ã€‚",
}
]


æ°¸é è¼¸å‡ºç‚º JSON é™£åˆ—ï¼Œä¸ä½¿ç”¨å·¢ç‹€JSONã€‚
"""
# âœ… Streamlit UI
st.title("ğŸ“‚ é‡‘å¡æª”æ¡ˆ OCR â†’ Excel è½‰æ›å·¥å…· ")

uploaded_files = st.file_uploader(
    "è«‹ä¸Šå‚³å¤šå€‹ PDF æˆ–å½±åƒæª” (PDF, PNG, JPG)", 
    type=["pdf", "png", "jpg", "jpeg"], 
    accept_multiple_files=True
)

if uploaded_files:
    all_records = []

    with st.spinner("ğŸš€ æ­£åœ¨è™•ç†æ‰€æœ‰æª”æ¡ˆï¼Œè«‹ç¨å€™..."):
        for file in uploaded_files:
            st.write(f"ğŸ” è™•ç†æª”æ¡ˆï¼š**{file.name}** ...")
            try:
                uploaded = genai.upload_file(file, mime_type=file.type)
                response = model.generate_content([OCR_PROMPT, uploaded])
                ocr_text = response.text.strip()
                ocr_text = re.sub(r"^```json\s*", "", ocr_text)
                ocr_text = re.sub(r"```$", "", ocr_text)
                ocr_text = ocr_text.replace("\ufeff", "").strip()

                st.subheader(f"ğŸ“œ Gemini å›å‚³å…§å®¹ ({file.name})")
                st.code(ocr_text, language="json")

                if not ocr_text:
                    st.error(f"âŒ {file.name} OCR å¤±æ•—ï¼šGemini API æ²’æœ‰å›å‚³å…§å®¹")
                    continue

                try:
                    records = json.loads(ocr_text)
                    if isinstance(records, list):
                        for r in records:
                            r["ä¾†æºæª”æ¡ˆ"] = file.name
                        all_records.extend(records)
                        st.success(f"âœ… {file.name} è§£æå®Œæˆï¼Œå…± {len(records)} ç­†è³‡æ–™")
                    else:
                        st.warning(f"âš ï¸ {file.name} çš„å›å‚³ä¸æ˜¯ JSON é™£åˆ—æ ¼å¼")
                except json.JSONDecodeError:
                    st.error(f"âŒ {file.name} OCR å›å‚³ä¸æ˜¯æœ‰æ•ˆçš„ JSONï¼Œè«‹æª¢æŸ¥ä¸Šæ–¹å…§å®¹")
            except Exception as e:
                st.error(f"âŒ {file.name} è™•ç†å¤±æ•—ï¼š{e}")

    # åŒ¯ç¸½çµæœè½‰ Excel
    # åŒ¯ç¸½çµæœè½‰ Excel
    if all_records:
        df = pd.DataFrame(all_records)
        df = df[[col for col in COLUMN_ORDER if col in df.columns] + ["ä¾†æºæª”æ¡ˆ"]]
        st.success(f"ğŸ‰ æ‰€æœ‰æª”æ¡ˆå®Œæˆï¼Œå…±è§£æ {len(df)} ç­†è³‡æ–™")

        # ç”¢ç”Ÿ Excel æª”æ¡ˆï¼ˆå…ˆå¯«å…¥ï¼‰
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        # ğŸ”§ ä½¿ç”¨ openpyxl é€²è¡Œç¾åŒ–ï¼šå›ºå®šæ¬„ä½å¯¬åº¦ã€æ¨™é¡Œé¡è‰²ã€å­—å‹èˆ‡æ›è¡Œ
        wb = load_workbook(buffer)
        ws = wb.active
        default_font = Font(size=14)

        # 1ï¸âƒ£ æ¨™é¡Œåˆ—ç¾åŒ–
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2ï¸âƒ£ æ¬„ä½å¯¬åº¦è¨­å®š
        width_map = {
               "è‹±æ–‡åå­—ï¼‹è‹±æ–‡å§“æ°": 27,
                "ä¸­æ–‡å§“å": 14,
                "ç¾è·å…¬å¸": 28,
                "ç¾è·è·ç¨±": 24,
                "æ•™è‚²èƒŒæ™¯(å­¸æ ¡)": 28,
                "æ•™è‚²èƒŒæ™¯(ç³»æ‰€)": 30,
                "å·¥ä½œç¶“æ­·": 80,         # æ–‡å­—å¤š â†’ åŠ å¯¬ + wrap_text
                "å°ˆæ¥­æŠ€è¡“/è­‰ç…§": 40,     # å¯èƒ½å¤šç­†
                "å°ˆé•·é ˜åŸŸ": 28,
                "å¯©æŸ¥æ„è¦‹æˆ–å‚™æ³¨": 90
        }

        # å¥—ç”¨æ¬„ä½å¯¬åº¦
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name in width_map:
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = width_map[col_name]

        # 3ï¸âƒ£ è¨­å®šåˆ—é«˜èˆ‡æ–‡å­—æ ¼å¼ï¼ˆå…¨æ¬„çµ±ä¸€ï¼‰
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 180  # å›ºå®šé«˜åº¦ï¼Œé¿å…æ›è¡Œè¢«æ“ å£“

        # 4ï¸âƒ£ å…¨æ¬„å¥—ç”¨å­—é«”ã€æ›è¡Œã€é ä¸Šå°é½Š
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    cell.font = default_font
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # å„²å­˜å›ç·©è¡å€
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        

        # æä¾›ä¸‹è¼‰
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="â¬‡ï¸ ä¸‹è¼‰åŒ¯ç¸½ Excel",
            data=buffer,
            file_name=f"OCRçµæœ_{now}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # âœ… Excel é è¦½
        st.subheader("ğŸ“Š Excel é è¦½")
        st.dataframe(df, use_container_width=True)