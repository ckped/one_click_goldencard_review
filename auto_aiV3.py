import os
import io
import json
import pandas as pd
import streamlit as st
import google.generativeai as genai
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# âœ… è®€å–ç’°å¢ƒè®Šæ•¸
from dotenv import load_dotenv
load_dotenv()
API_KEY = os.getenv("GENIMI_API_KEY")
if not API_KEY:
    st.error("âŒ æ‰¾ä¸åˆ°ç’°å¢ƒè®Šæ•¸ GENIMI_API_KEYï¼Œè«‹å…ˆè¨­å®šå¾Œé‡æ–°åŸ·è¡Œ")
else:
    genai.configure(api_key=API_KEY)

# âœ… æ¨¡å‹èˆ‡æ¬„ä½è¨­å®š
model = genai.GenerativeModel("gemini-2.5-pro")
COLUMN_ORDER = [
    "è‹±æ–‡åå­—ï¼‹è‹±æ–‡å§“æ°", "ä¸­æ–‡å§“å", "æ€§åˆ¥", "åœ‹ç±", "å­¸æ­·ï¼‹å­¸æ ¡ï¼‹ç§‘ç³»",
    "ç”³è«‹è³‡æ ¼æ¢ä»¶", "å‡ºç”Ÿæ—¥æœŸ", "è­·ç…§è™Ÿç¢¼", "å­é ˜åŸŸ", "ç¾è·å…¬å¸", "ç¾è·è·ç¨±",
    "å…¶ä»–å·¥ä½œç¶“æ­·", "ç¾è·æ˜¯å¦ç‚ºä¸»ç®¡", "æ•™è‚²èƒŒæ™¯(å­¸æ ¡)", "æ•™è‚²èƒŒæ™¯(ç³»æ‰€)",
    "å·¥ä½œç¶“æ­·", "ç”¢æ¥­å¯¦ç¸¾å°ˆé•·", "æ±‚å­¸æœŸé–“", "ç•¢æ¥­å¹´ä»½", "ç¸½å·¥ä½œå¹´è³‡",
    "å¯©æŸ¥æ„è¦‹æˆ–å‚™æ³¨", "å‹å‹•éƒ¨æª¢æ ¸çµæœ", "æ˜¯å¦ç‚ºè½‰è‡ªå…¶ä»–é ˜åŸŸ", "ç¬¬1æ¬¡ç”³è«‹", "å¹´é½¡", "æœˆè–ª"
]

OCR_PROMPT = """è§’è‰²ï¼šä½ æ˜¯ç²¾æº–çš„è³‡æ–™åˆ†æåŠ©ç†ï¼Œå°ˆè²¬å¾ç”³è«‹äººè³‡æ–™ PDF ä¸­æå–é—œéµè³‡è¨Šã€‚

ä»»å‹™ï¼š
1. è§£æ PDFï¼Œå…§å®¹å«ç”³è«‹æ›¸ã€è­·ç…§ã€å­¸æ­·ã€ç¶“æ­·èˆ‡è–ªè³‡è­‰æ˜ï¼Œå¯èƒ½å«ä¸­æ–‡ã€è‹±æ–‡æˆ–å…¶ä»–å¤–æ–‡ã€‚
2. å°æ¯ä½ç”³è«‹äººå–®ç¨è¼¸å‡º JSONï¼ˆå¤šä½å‰‡ä¾åºç”Ÿæˆå¤šç­†ï¼‰ã€‚
3. æ‰€æœ‰æ¬„ä½åç¨±é ˆèˆ‡æŒ‡å®šå®Œå…¨ä¸€è‡´ï¼Œç¼ºæ¼å€¼å¡« "N/A"ï¼Œé‡‘é¡é ˆåŒ…å«å¹£åˆ¥ã€‚

å·¥ä½œç¶“æ­·åˆ¤æ–·ï¼š
- è‹¥ç”³è«‹ç‚ºã€Œå…·æ•¸ä½ç¶“æ¿Ÿç›¸é—œç”¢æ¥­ã€8å¹´ä»¥ä¸Šç¶“é©—ã€ï¼š
  1. åˆ¤æ–·å…¬å¸æ˜¯å¦å±¬è»Ÿé«”/è³‡è¨Šæœå‹™æ¥­ï¼ˆå¦‚è»Ÿé«”é–‹ç™¼ã€ç³»çµ±æ•´åˆã€é›²ç«¯æœå‹™ï¼‰ã€‚
  2. éè»Ÿé«”æ¥­æ™‚ï¼Œé ˆæª¢è¦–ç”³è«‹äººã€Œè·å‹™å…§å®¹ã€æ˜¯å¦ç¬¦åˆè»Ÿé«”é–‹ç™¼ã€ç³»çµ±æ¶æ§‹æˆ–è³‡è¨ŠæŠ€è¡“æ ¸å¿ƒè·èƒ½ã€‚
  3. æ•™è‚²/åŸ¹è¨“é ˜åŸŸä¸è¦–ç‚ºæ•¸ä½ç”¢æ¥­ï¼Œåƒ…åˆ—ç‚ºè¼”åŠ©ç¶“æ­·ã€‚è‹¥æè¿°éæ–¼ç°¡ç•¥ï¼Œé ˆè¦æ±‚è£œå……å°ˆæ¡ˆèˆ‡è²¢ç»ã€‚

è¼¸å‡ºæ ¼å¼ï¼ˆåƒ…è¼¸å‡º JSON å­—ä¸²ï¼‰ï¼š
[
  {
    "è‹±æ–‡åå­—ï¼‹è‹±æ–‡å§“æ°": "string",
    "ä¸­æ–‡å§“å": "string",
    "æ€§åˆ¥": "string",
    "åœ‹ç±": "string",
    "å­¸æ­·ï¼‹å­¸æ ¡ï¼‹ç§‘ç³»": "stringï¼Œä¾å¤§å­¸ã€ç ”ç©¶æ‰€ã€åšå£«é †åºåˆ†é»",
    "ç”³è«‹è³‡æ ¼æ¢ä»¶": "string",
    "å‡ºç”Ÿæ—¥æœŸ": "YYYY-MM-DD",
    "è­·ç…§è™Ÿç¢¼": "string",
    "å­é ˜åŸŸ": "string",
    "ç¾è·å…¬å¸": "string",
    "ç¾è·è·ç¨±": "string",
    "å…¶ä»–å·¥ä½œç¶“æ­·": "string",
    "ç¾è·æ˜¯å¦ç‚ºä¸»ç®¡": "string",
    "æ•™è‚²èƒŒæ™¯(å­¸æ ¡)": "stringï¼Œä¾å¤§å­¸ã€ç ”ç©¶æ‰€ã€åšå£«é †åºï¼Œæ ¼å¼ï¼šå¤§å­¸ï¼šæ ¡åã€ç ”ç©¶æ‰€ï¼šæ ¡å",
    "æ•™è‚²èƒŒæ™¯(ç³»æ‰€)": "stringï¼Œä¾å¤§å­¸ã€ç ”ç©¶æ‰€ã€åšå£«é †åºï¼Œæ ¼å¼ï¼šå¤§å­¸ï¼šç³»æ‰€ã€ç ”ç©¶æ‰€ï¼šç³»æ‰€",
    "å·¥ä½œç¶“æ­·": "stringï¼Œåˆ—å…¬å¸ã€è·ç¨±ã€æœŸé–“(YYYY-MM)ã€å·¥ä½œæè¿°èˆ‡è–ªè³‡ï¼ˆå«å¹£åˆ¥ï¼‰",
    "ç”¢æ¥­å¯¦ç¸¾å°ˆé•·": "string",
    "æ±‚å­¸æœŸé–“": "string",
    "ç•¢æ¥­å¹´ä»½": "æ•¸å­—",
    "ç¸½å·¥ä½œå¹´è³‡": "æ•¸å­—ï¼Œå–®ä½å¹´",
    "å¯©æŸ¥æ„è¦‹æˆ–å‚™æ³¨": "stringï¼Œå…ˆåˆ¤å®šã€Œé€šé/æœªé€šé/å¾…ç¢ºèªã€ï¼Œå†è©³è¿°è©•ä¼°èˆ‡ç–‘é»ï¼ˆå¦‚éè»Ÿé«”æ¥­ã€åƒ…æ•¸ä½å·¥å…·ä½¿ç”¨ã€æˆ–æè¿°ä¸è¶³ï¼‰ã€‚",
    "å‹å‹•éƒ¨æª¢æ ¸çµæœ": "string",
    "æ˜¯å¦ç‚ºè½‰è‡ªå…¶ä»–é ˜åŸŸ": "string",
    "ç¬¬1æ¬¡ç”³è«‹": "string",
    "å¹´é½¡": "string",
    "æœˆè–ª": "stringï¼Œå«å¹£åˆ¥èˆ‡é‡‘é¡ï¼ˆä¾‹ï¼šUSD 3000ï¼Œé™„å°å¹£æ›ç®—ï¼‰"
  }
]
"""
# âœ… Streamlit UI
st.title("ğŸ“‚ é‡‘å¡æª”æ¡ˆ OCR â†’ Excel è½‰æ›å·¥å…· ")

uploaded_files = st.file_uploader(
    "è«‹ä¸Šå‚³å¤šå€‹ PDF æˆ–å½±åƒæª” (PDF, PNG, JPG)", 
    type=["pdf", "png", "jpg", "jpeg"], 
    accept_multiple_files=True
)

if uploaded_files:
    all_records = []

    with st.spinner("ğŸš€ æ­£åœ¨è™•ç†æ‰€æœ‰æª”æ¡ˆï¼Œè«‹ç¨å€™..."):
        for file in uploaded_files:
            st.write(f"ğŸ” è™•ç†æª”æ¡ˆï¼š**{file.name}** ...")
            try:
                uploaded = genai.upload_file(file, mime_type=file.type)
                response = model.generate_content([OCR_PROMPT, uploaded])
                ocr_text = response.text.strip()
                ocr_text = re.sub(r"^```json\s*", "", ocr_text)
                ocr_text = re.sub(r"```$", "", ocr_text)
                ocr_text = ocr_text.replace("\ufeff", "").strip()

                st.subheader(f"ğŸ“œ Gemini å›å‚³å…§å®¹ ({file.name})")
                st.code(ocr_text, language="json")

                if not ocr_text:
                    st.error(f"âŒ {file.name} OCR å¤±æ•—ï¼šGemini API æ²’æœ‰å›å‚³å…§å®¹")
                    continue

                try:
                    records = json.loads(ocr_text)
                    if isinstance(records, list):
                        for r in records:
                            r["ä¾†æºæª”æ¡ˆ"] = file.name
                        all_records.extend(records)
                        st.success(f"âœ… {file.name} è§£æå®Œæˆï¼Œå…± {len(records)} ç­†è³‡æ–™")
                    else:
                        st.warning(f"âš ï¸ {file.name} çš„å›å‚³ä¸æ˜¯ JSON é™£åˆ—æ ¼å¼")
                except json.JSONDecodeError:
                    st.error(f"âŒ {file.name} OCR å›å‚³ä¸æ˜¯æœ‰æ•ˆçš„ JSONï¼Œè«‹æª¢æŸ¥ä¸Šæ–¹å…§å®¹")
            except Exception as e:
                st.error(f"âŒ {file.name} è™•ç†å¤±æ•—ï¼š{e}")

    # åŒ¯ç¸½çµæœè½‰ Excel
    # åŒ¯ç¸½çµæœè½‰ Excel
    if all_records:
        df = pd.DataFrame(all_records)
        df = df[[col for col in COLUMN_ORDER if col in df.columns] + ["ä¾†æºæª”æ¡ˆ"]]
        st.success(f"ğŸ‰ æ‰€æœ‰æª”æ¡ˆå®Œæˆï¼Œå…±è§£æ {len(df)} ç­†è³‡æ–™")

        # ç”¢ç”Ÿ Excel æª”æ¡ˆï¼ˆå…ˆå¯«å…¥ï¼‰
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        # ğŸ”§ ä½¿ç”¨ openpyxl é€²è¡Œç¾åŒ–ï¼šå›ºå®šæ¬„ä½å¯¬åº¦ã€æ¨™é¡Œé¡è‰²ã€å­—å‹èˆ‡æ›è¡Œ
        wb = load_workbook(buffer)
        ws = wb.active
        default_font = Font(size=14)

        # 1ï¸âƒ£ æ¨™é¡Œåˆ—ç¾åŒ–
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2ï¸âƒ£ æ¬„ä½å¯¬åº¦è¨­å®š
        width_map = {
            "è‹±æ–‡åå­—ï¼‹è‹±æ–‡å§“æ°": 25,
            "ä¸­æ–‡å§“å": 12,
            "æ€§åˆ¥": 10,
            "åœ‹ç±": 12,
            "å­¸æ­·ï¼‹å­¸æ ¡ï¼‹ç§‘ç³»": 30,
            "ç”³è«‹è³‡æ ¼æ¢ä»¶": 25,
            "å‡ºç”Ÿæ—¥æœŸ": 15,
            "è­·ç…§è™Ÿç¢¼": 18,
            "å­é ˜åŸŸ": 20,
            "ç¾è·å…¬å¸": 25,
            "ç¾è·è·ç¨±": 20,
            "å…¶ä»–å·¥ä½œç¶“æ­·": 40,
            "ç¾è·æ˜¯å¦ç‚ºä¸»ç®¡": 15,
            "æ•™è‚²èƒŒæ™¯(å­¸æ ¡)": 35,
            "æ•™è‚²èƒŒæ™¯(ç³»æ‰€)": 35,
            "å·¥ä½œç¶“æ­·": 80,
            "ç”¢æ¥­å¯¦ç¸¾å°ˆé•·": 60,
            "æ±‚å­¸æœŸé–“": 20,
            "ç•¢æ¥­å¹´ä»½": 12,
            "ç¸½å·¥ä½œå¹´è³‡": 15,
            "å¯©æŸ¥æ„è¦‹æˆ–å‚™æ³¨": 80,
            "å‹å‹•éƒ¨æª¢æ ¸çµæœ": 40,
            "æ˜¯å¦ç‚ºè½‰è‡ªå…¶ä»–é ˜åŸŸ": 30,
            "ç¬¬1æ¬¡ç”³è«‹": 15,
            "å¹´é½¡": 10,
            "æœˆè–ª": 25,
            "ä¾†æºæª”æ¡ˆ": 30
        }

        # å¥—ç”¨æ¬„ä½å¯¬åº¦
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name in width_map:
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = width_map[col_name]

        # 3ï¸âƒ£ è¨­å®šåˆ—é«˜èˆ‡æ–‡å­—æ ¼å¼ï¼ˆå…¨æ¬„çµ±ä¸€ï¼‰
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 180  # å›ºå®šé«˜åº¦ï¼Œé¿å…æ›è¡Œè¢«æ“ å£“

        # 4ï¸âƒ£ å…¨æ¬„å¥—ç”¨å­—é«”ã€æ›è¡Œã€é ä¸Šå°é½Š
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    cell.font = default_font
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # å„²å­˜å›ç·©è¡å€
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        

        # æä¾›ä¸‹è¼‰
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="â¬‡ï¸ ä¸‹è¼‰åŒ¯ç¸½ Excel",
            data=buffer,
            file_name=f"OCRçµæœ_{now}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # âœ… Excel é è¦½
        st.subheader("ğŸ“Š Excel é è¦½")
        st.dataframe(df, use_container_width=True)